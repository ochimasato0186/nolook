"""
NO LOOK API (FastAPI)
- Endpoints: / (health), /ask, /analyze, /summary, /weekly_report, /export, /metrics
- DB: SQLAlchemy (SQLite by default)
- LLM: LangChain+OpenAI if available (hybrid classifier). 無効時は辞書ルールのみ。
- Security: API Key, simple rate limit, request logging + Prometheus metrics
"""
from __future__ import annotations

import json
import os
import re
import time
import logging
import uuid
import yaml
import io, csv, datetime as dt
import zoneinfo
from typing import Dict, List, Optional, Literal
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

from fastapi import FastAPI, HTTPException, Depends, Request, Security, Query, Body
from fastapi.security import APIKeyHeader
from fastapi.middleware.cors import CORSMiddleware
from fastapi.exceptions import RequestValidationError
from fastapi.responses import StreamingResponse, JSONResponse
from starlette import status
from pydantic import BaseModel
from dotenv import load_dotenv
from prometheus_client import (
    Counter, Histogram, Gauge, CollectorRegistry,
    generate_latest, CONTENT_TYPE_LATEST
)

# ============== Small helpers (export, formatting) ============================
def _fmt_created_local(dtobj: dt.datetime, tzname: str = "Asia/Tokyo") -> str:
    """UTC or naive datetime -> Asia/Tokyo の 'YYYY-MM-DD HH:MM' へ"""
    if dtobj is None:
        return ""
    if dtobj.tzinfo is None:
        dtobj = dtobj.replace(tzinfo=dt.timezone.utc)
    try:
        tzinfo = zoneinfo.ZoneInfo(tzname)
    except Exception:
        tzinfo = dt.timezone(dt.timedelta(hours=9))  # Fallback JST
    local = dtobj.astimezone(tzinfo)
    return local.strftime("%Y-%m-%d %H:%M")

def _class_display(v: Optional[str]) -> str:
    return v if (v is not None and str(v).strip() != "") else "-"

def _topics_join(tags: Optional[List[str]]) -> str:
    return ";".join(map(str, tags or []))

# ============== Env ===========================================================
load_dotenv()
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
DISABLE_OPENAI = os.getenv("NOLOOK_DISABLE_OPENAI", "0") == "1"
LLM_WEIGHT = float(os.getenv("NOLOOK_LLM_WEIGHT", "0.7") or 0.7)
API_KEY = os.getenv("API_KEY")
RATE_PER_MIN = int(os.getenv("NOLOOK_RATE_LIMIT_PER_MIN", "0") or 0)
ALLOWED_ORIGINS = [o.strip() for o in os.getenv(
    "ALLOWED_ORIGINS",
    "http://localhost:3000,http://127.0.0.1:3000"
).split(",") if o.strip()]

# 手動モード
MANUAL_ONLY = os.getenv("NOLOOK_MANUAL_ONLY", "0") == "1"

# Optional LangChain (Runnable仕様)
LANGCHAIN_AVAILABLE = False
try:
    if not DISABLE_OPENAI and OPENAI_API_KEY:
        from langchain_openai import ChatOpenAI
        from langchain_core.prompts import ChatPromptTemplate
        from langchain_core.output_parsers import StrOutputParser
        LANGCHAIN_AVAILABLE = True
except Exception:
    LANGCHAIN_AVAILABLE = False

# ============== DB: SQLAlchemy ===============================================
from sqlalchemy import (
    create_engine, Column, Integer, String, Float, Boolean, DateTime, JSON, func, select, and_
)
from sqlalchemy.orm import declarative_base, sessionmaker

DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./nolook_dev.db")
connect_args = {"check_same_thread": False} if DATABASE_URL.startswith("sqlite") else {}
engine = create_engine(DATABASE_URL, connect_args=connect_args, future=True)
SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
Base = declarative_base()

# 固定キー
EMOTION_KEYS = ["楽しい", "悲しい", "怒り", "不安", "しんどい", "中立"]

class StatsRecord(Base):
    __tablename__ = "stats_records"
    id = Column(Integer, primary_key=True, autoincrement=True)
    created_at = Column(DateTime(timezone=True), server_default=func.now(), nullable=False)
    class_id = Column(String, nullable=True)
    emotion = Column(String, nullable=False)
    score = Column(Float, nullable=False)
    labels = Column(JSON, nullable=False)
    topic_tags = Column(JSON, nullable=False)
    relationship_mention = Column(Boolean, nullable=False)
    negation_index = Column(Float, nullable=False)
    avoidance = Column(Float, nullable=False)

Base.metadata.create_all(bind=engine)

with engine.begin() as conn:
    try:
        conn.exec_driver_sql("CREATE INDEX IF NOT EXISTS idx_stats_created_at ON stats_records (created_at)")
        conn.exec_driver_sql("CREATE INDEX IF NOT EXISTS idx_stats_emotion ON stats_records (emotion)")
        conn.exec_driver_sql("CREATE INDEX IF NOT EXISTS idx_stats_class_id ON stats_records (class_id)")
    except Exception:
        pass

# ============== Logging & Metrics ============================================
logger = logging.getLogger("nolook")
if not logger.handlers:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

PROM_REG = CollectorRegistry()
REQ_COUNT = Counter("app_requests_total", "Total HTTP requests", ["path", "method", "status"], registry=PROM_REG)
REQ_LATENCY = Histogram("app_request_duration_seconds", "Request duration seconds", ["path"], registry=PROM_REG,
                        buckets=(0.01, 0.025, 0.05, 0.1, 0.25, 0.5, 1, 2, 5, 10))
EXC_COUNT = Counter("app_exceptions_total", "Exceptions by path", ["path"], registry=PROM_REG)
ANALYZE_COUNT = Counter("nolook_analyze_total", "Analyze saved by primary emotion", ["emotion"], registry=PROM_REG)
EXPORT_BYTES = Counter("nolook_export_bytes_total", "Exported bytes by format", ["format"], registry=PROM_REG)
WEEKLY_CACHE_HITS = Counter("nolook_weekly_report_cache_hits_total", "Weekly report cache hits", registry=PROM_REG)
WEEKLY_CACHE_MISSES = Counter("nolook_weekly_report_cache_misses_total", "Weekly report cache misses", registry=PROM_REG)
GAUGE_MANUAL = Gauge("nolook_manual_only", "Manual-only mode (0/1)", registry=PROM_REG)
GAUGE_LANGCHAIN = Gauge("nolook_langchain_available", "LangChain available (0/1)", registry=PROM_REG)
GAUGE_MANUAL.set(1.0 if (os.getenv("NOLOOK_MANUAL_ONLY","0")=="1") else 0.0)
GAUGE_LANGCHAIN.set(1.0 if LANGCHAIN_AVAILABLE else 0.0)

# ============== Utils =========================================================
def _full_labels(labels: dict) -> dict:
    return {k: float(labels.get(k, 0.0)) for k in EMOTION_KEYS}

def _normalize(labels: Dict[str, float]) -> Dict[str, float]:
    total = float(sum(max(0.0, v) for v in labels.values()))
    if total <= 0.0:
        return {k: (1.0 if k == "中立" else 0.0) for k in EMOTION_KEYS}
    return {k: round(max(0.0, labels.get(k, 0.0)) / total, 4) for k in EMOTION_KEYS}

def _one_hot(emotion: str) -> Dict[str, float]:
    return {k: (1.0 if k == emotion else 0.0) for k in EMOTION_KEYS}

# ============== Emotion alias =================================================
DEFAULT_EMO_ALIASES = {
    "楽しみ": "楽しい",
    "たのしい": "楽しい",
    "うれしい": "楽しい",
    "普通": "中立",
    "ふつう": "中立",
    "ok": "中立",
    "OK": "中立",
    "ムカつく": "怒り",
    "むかつく": "怒り",
    "こわい": "不安",
    "怖い": "不安",
    "疲れた": "しんどい",
}
def _width_fold(s: str) -> str:
    try:
        import unicodedata
        s = unicodedata.normalize("NFKC", s)
    except Exception:
        pass
    return s.casefold().strip()

def _load_alias_map() -> Dict[str, str]:
    alias: Dict[str, str] = {_width_fold(k): v for k, v in DEFAULT_EMO_ALIASES.items()}
    try:
        yaml_path = (Path(__file__).resolve().parent.parent / "config" / "emotion_aliases.yaml")
        if yaml_path.exists():
            with yaml_path.open("r", encoding="utf-8") as f:
                data = yaml.safe_load(f) or {}
            for canonical, aliases in (data or {}).items():
                alias[_width_fold(canonical)] = canonical
                for a in aliases or []:
                    alias[_width_fold(str(a))] = canonical
    except Exception as e:
        logger.warning("failed to load emotion_aliases.yaml: %s", e)
    for k in EMOTION_KEYS:
        alias[_width_fold(k)] = k
    return alias

ALIAS_MAP = _load_alias_map()
for _k in EMOTION_KEYS:
    ALIAS_MAP.setdefault(_k, _k)
    ALIAS_MAP.setdefault(_width_fold(_k), _k)

def normalize_emotion_name(name: Optional[str]) -> Optional[str]:
    if not name:
        return None
    n = _width_fold(str(name))
    for k in EMOTION_KEYS:
        if _width_fold(k) == n:
            return k
    if n in ALIAS_MAP:
        return ALIAS_MAP[n]
    for k in EMOTION_KEYS:
        if _width_fold(k) in n or n in _width_fold(k):
            return k
    return None

# ============== Lexicon classification =======================================
EMOTION_LEXICON: Dict[str, List[str]] = {
    "楽しい": ["楽しい", "嬉しい", "うれしい", "わくわく", "最高", "やった",
             "自己ベスト", "満点", "合格", "優勝", "成功", "勝った"],
    "悲しい": ["悲しい", "さみしい", "辛い", "つらい", "泣きたい", "萎え", "凹む"],
    "怒り":   ["怒", "ムカつ", "むかつ", "腹立", "イライラ", "キレた"],
    "不安":   ["不安", "心配", "怖い", "こわい", "緊張", "心細い"],
    "しんどい": ["疲れ", "だる", "しんど", "眠い", "疲労"],
}
JOY_CONTEXT_PATTERNS = [
    r"(テスト|模試|試験|結果|成績).*(最高|自己ベスト|満点|できた|更新|良かった)",
    r"(合格|受かった|当選|優勝|入賞|表彰|MVP)",
]
NEG_NEAR_PATTERN = re.compile(r"(じゃない|じゃなかった|なくはない|ない|なかった)")

def _apply_context_boosts(text: str, counts: Dict[str, int]) -> None:
    t = text or ""
    if NEG_NEAR_PATTERN.search(t):
        return
    if any(re.search(p, t) for p in JOY_CONTEXT_PATTERNS):
        counts["楽しい"] += 2

def _lexicon_dist(text: str) -> Dict[str, float]:
    t = text or ""
    counts = defaultdict(int)
    for emo, kws in EMOTION_LEXICON.items():
        for kw in kws:
            counts[emo] += len(re.findall(re.escape(kw), t))
    _apply_context_boosts(t, counts)
    if sum(counts.values()) == 0:
        return {"中立": 1.0}
    labels = {emo: float(c) for emo, c in counts.items()}
    return _normalize(_full_labels(labels))

# ============== LangChain (Runnable版) ========================================
classification_chain = None
if LANGCHAIN_AVAILABLE:
    _cls_prompt = ChatPromptTemplate.from_messages([
        ("system", "あなたは中高生の日本語日記の感情分類器です。"
                   "6感情 [\"楽しい\",\"悲しい\",\"怒り\",\"不安\",\"しんどい\",\"中立\"] の分布をJSONだけで返してください。"
                   "例: {\"labels\":{\"楽しい\":0.1,\"悲しい\":0.6,\"怒り\":0.1,\"不安\":0.1,\"しんどい\":0.1,\"中立\":0.0},\"primary\":\"悲しい\"}"),
        ("human", "テキスト: {text}")
    ])
    _cls_llm = ChatOpenAI(model="gpt-4o-mini", temperature=0.2)
    classification_chain = _cls_prompt | _cls_llm | StrOutputParser()

def _coerce_text(x) -> str:
    if isinstance(x, str):
        return x
    c = getattr(x, "content", None)
    if isinstance(c, str):
        return c
    try:
        return json.dumps(x, ensure_ascii=False)
    except Exception:
        return str(x)

def _llm_dist(text: str) -> Optional[Dict[str, float]]:
    if not classification_chain:
        return None
    try:
        raw = classification_chain.invoke({"text": text})
        raw_text = _coerce_text(raw)
        data = json.loads(raw_text)
        labels = data.get("labels", {})
        return _normalize(_full_labels(labels))
    except Exception:
        return None

def classify_emotion(text: str):
    lex = _lexicon_dist(text)
    llm = _llm_dist(text) if classification_chain else None
    if llm:
        combined = {k: (1.0 - LLM_WEIGHT) * lex.get(k, 0.0) + LLM_WEIGHT * llm.get(k, 0.0) for k in EMOTION_KEYS}
        labels = _normalize(combined)
    else:
        labels = _normalize(lex)
    top_emotion = max(labels.items(), key=lambda x: x[1])[0]
    top_score = float(labels[top_emotion])
    return top_emotion, top_score, labels

def resolve_emotion(text: str, selected_emotion: Optional[str]):
    norm = normalize_emotion_name(selected_emotion)
    if norm:
        return norm, 1.0, _one_hot(norm)
    manual_only = (os.getenv("NOLOOK_MANUAL_ONLY", "0") == "1") or bool(globals().get("MANUAL_ONLY", False))
    if manual_only:
        raise HTTPException(status_code=400, detail="selected_emotion を指定してください（例: 楽しい/悲しい/怒り/不安/しんどい/中立）")
    return classify_emotion(text)

# ============== Signals ======================================================
class Signals(BaseModel):
    topic_tags: List[str]
    relationship_mention: bool
    negation_index: float
    avoidance: float

TOPIC_LEXICON: Dict[str, List[str]] = {
    "友だち": ["友だち", "友達", "ともだち", "いじめ", "無視", "仲間", "先輩", "後輩"],
    "勉強": ["勉強", "テスト", "宿題", "成績", "授業", "課題", "受験"],
    "家庭": ["家", "家族", "親", "父", "母", "兄", "姉", "弟", "妹"],
    "部活": ["部活", "クラブ", "サークル", "試合", "大会", "練習"],
    "体調": ["体調", "熱", "風邪", "腹痛", "頭痛", "眠い", "疲れ", "しんどい"],
}
RELATIONSHIP_WORDS = ["友だち", "友達", "ともだち", "いじめ", "無視", "悪口", "仲間はずれ", "ぼっち"]
NEGATION_WORDS = ["ない", "できない", "無理", "嫌い", "いやだ", "ダメ", "もうやだ"]
AVOIDANCE_WORDS = ["別に", "なんでもない", "知らない", "まあいい", "どうでもいい"]

def compute_signals(text: str) -> Signals:
    t = text or ""
    tags: List[str] = []
    for tag, kws in TOPIC_LEXICON.items():
        if any(kw in t for kw in kws):
            tags.append(tag)
    rel = any(w in t for w in RELATIONSHIP_WORDS)

    def score_by_terms(terms: List[str]) -> float:
        hits = sum(t.count(w) for w in terms)
        return float(min(1.0, round(hits / 10.0, 2)))

    return Signals(
        topic_tags=tags,
        relationship_mention=rel,
        negation_index=score_by_terms(NEGATION_WORDS),
        avoidance=score_by_terms(AVOIDANCE_WORDS),
    )

# ============== Replies (Runnable版) =========================================
PHRASE_LIB: Dict[str, List[str]] = {
    "楽しい": ["楽しさが伝わってきて、こっちまで笑顔になったよ。", "がんばりが実ってよかったね。"],
    "悲しい": ["それは本当にこたえたね…。", "気持ちを言葉にしてくれて、まずはえらいよ。"],
    "怒り":   ["それは腹が立つよね、わかるよ。", "気持ちを守るために距離を取るのもありだよ。"],
    "不安":   ["不安な気持ち、ちゃんと伝わってきたよ。", "いまの君なりのペースで大丈夫。"],
    "しんどい": ["しんどさ、よくがんばって耐えてきたね。", "今日は自分を少し甘やかしてあげよう。"],
    "中立": ["話してくれてありがとう。", "状況はわかったよ。"],
}
FOLLOWUP_TAILS: Dict[str, str] = {
    "楽しい": "一番『よかった！』って思えた瞬間はどこだった？",
    "悲しい": "いま心が少し軽くなるとしたら、どんな声かけが嬉しい？",
    "怒り":   "少し落ち着いたら、どう動くのが自分らしいと思う？",
    "不安":   "少しでもいいから、いっしょにやれること決めてみる？",
    "しんどい": "今日はどんなふうに休めそうかな？無理なくいこう。",
    "中立":   "他にも感じたことがあれば、いつでも聞かせてね。",
}
def _tail_for(emotion: str) -> str:
    return FOLLOWUP_TAILS.get(emotion, "何か話したいことがあれば、いつでも聞かせてね。")

def _join_short(lines: List[str], max_sent: int) -> str:
    msg = " ".join(lines[:max_sent])
    return msg.strip()

reply_chain = None
if LANGCHAIN_AVAILABLE:
    _reply_prompt = ChatPromptTemplate.from_messages([
        ("system",
         "日本語で返答。口調: {style}（buddy=友だち風/coach=前向き/teacher=落ち着いた先生）。"
         "絵文字・箇条書きなし。命令調で宿題のような指示はしない。"
         "長さ: {length}（medium=最大2文、1文は60字以内）。"
         "必ず含める: ユーザー内容への短い共感/ささやかな励まし/可能なら topics: {topics} をさりげなく踏まえる"),
        ("human", "{text}")
    ])
    _reply_llm = ChatOpenAI(model="gpt-4o-mini", temperature=0.4)
    reply_chain = _reply_prompt | _reply_llm | StrOutputParser()

def build_reply(text: str, emotion: str, sig, *, style: str = "buddy", length: str = "medium", followup: bool = False) -> str:
    if reply_chain is not None:
        try:
            base_raw = reply_chain.invoke({
                "text": text, "emotion": emotion, "style": style,
                "length": length, "topics": ",".join(getattr(sig, "topic_tags", []))
            })
            base = _coerce_text(base_raw)
            return f"{base} {_tail_for(emotion)}" if followup else base
        except Exception:
            pass
    base_lines = PHRASE_LIB.get(emotion, PHRASE_LIB["中立"]).copy()
    msg = _join_short(base_lines, 2)
    return f"{msg} {_tail_for(emotion)}" if followup else msg

# （この後は save_stats, Security, FastAPI app, Endpoints, Summary, Weekly Report, Export, __main__ が続きます）

# ============== 永続化 ========================================================
class _SignalsLike:
    topic_tags: List[str]
    relationship_mention: bool
    negation_index: float
    avoidance: float

def save_stats(*, labels: dict, signals: _SignalsLike, top_emotion: str, score: float, class_id: Optional[str] = None) -> int:
    rec = StatsRecord(
        class_id=class_id,
        emotion=top_emotion,
        score=score,
        labels=_full_labels(labels),
        topic_tags=list(getattr(signals, "topic_tags", [])),
        relationship_mention=bool(getattr(signals, "relationship_mention", False)),
        negation_index=float(getattr(signals, "negation_index", 0.0)),
        avoidance=float(getattr(signals, "avoidance", 0.0)),
    )
    with SessionLocal() as db:
        db.add(rec)
        db.commit()
        db.refresh(rec)
        return rec.id

# ============== Security (API Key & Rate Limit) ==============================
api_key_header = APIKeyHeader(name="X-API-Key", auto_error=False)

async def verify_api_key(api_key: Optional[str] = Security(api_key_header)):
    if not API_KEY:
        return True
    if not api_key:
        raise HTTPException(status_code=401, detail="Missing API key")
    if api_key != API_KEY:
        raise HTTPException(status_code=403, detail="Invalid API key")
    return True

RATE_STORE: Dict[str, tuple] = {}

async def rate_limit(req: Request):
    if RATE_PER_MIN <= 0:
        return True
    key = f"{req.client.host}:{req.url.path}"
    now = int(time.time())
    window = now // 60
    cur = RATE_STORE.get(key)
    if not cur or cur[0] != window:
        RATE_STORE[key] = (window, 1)
        return True
    cnt = cur[1] + 1
    RATE_STORE[key] = (window, cnt)
    if cnt > RATE_PER_MIN:
        raise HTTPException(status_code=429, detail="Rate limit exceeded")
    return True

# ============== FastAPI app ===================================================
app = FastAPI(title="NO LOOK API", version="0.5.3")

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST", "OPTIONS"],
    allow_headers=["*"],
    max_age=3600,
)

# Request-ID + ログ
@app.middleware("http")
async def add_request_id_and_log(request: Request, call_next):
    rid = str(uuid.uuid4())
    request.state.rid = rid
    start = time.time()
    try:
        response = await call_next(request)
    except Exception as e:
        logger.exception("rid=%s unhandled error: %s", rid, e)
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"error": {"type": "internal_error", "message": "Internal Server Error", "rid": rid}},
        )
    dur = int((time.time() - start) * 1000)
    response.headers["X-Request-ID"] = rid
    logger.info("rid=%s method=%s path=%s status=%s dur_ms=%s", rid, request.method, request.url.path, response.status_code, dur)
    return response

# メトリクス
@app.middleware("http")
async def metrics_middleware(request: Request, call_next):
    path = request.url.path
    start_t = time.perf_counter()
    try:
        response = await call_next(request)
    except Exception:
        EXC_COUNT.labels(path=path).inc()
        raise
    finally:
        REQ_LATENCY.labels(path=path).observe(time.perf_counter() - start_t)
    REQ_COUNT.labels(path=path, method=request.method, status=str(getattr(response, "status_code", 500))).inc()
    return response

# バリデーションエラー
@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    rid = getattr(request.state, "rid", str(uuid.uuid4()))
    return JSONResponse(
        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
        content={"error": {"type": "validation_error", "message": exc.errors(), "rid": rid}},
    )

# ============== Schemas（/ask 用のみ） =======================================
class AskRequest(BaseModel):
    prompt: str
    selected_emotion: Optional[str] = None
    style: Optional[Literal["buddy", "coach", "teacher"]] = "buddy"
    length: Optional[Literal["medium"]] = "medium"
    followup: Optional[bool] = False

class AskResponse(BaseModel):
    reply: str
    emotion: str
    score: float
    labels: Dict[str, float]

class AnalyzeResponse(BaseModel):
    labels: Dict[str, float]
    signals: Signals

# ============== Endpoints =====================================================
@app.get("/")
async def read_root(verify: bool = Depends(verify_api_key)):
    return {
        "message": "NO LOOK API running",
        "openai": LANGCHAIN_AVAILABLE,
        "llm_weight": LLM_WEIGHT,
        "manual_only": ((os.getenv("NOLOOK_MANUAL_ONLY", "0") == "1") or bool(globals().get("MANUAL_ONLY", False))),
    }

@app.post("/ask", response_model=AskResponse)
async def ask_ai(request: AskRequest, _k=Depends(verify_api_key), _r=Depends(rate_limit)):
    text = (request.prompt or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="prompt が空です。")
    emotion, score, labels = resolve_emotion(text, request.selected_emotion)
    sig = compute_signals(text)
    reply = build_reply(text, emotion, sig, style=request.style or "buddy", length="medium", followup=bool(request.followup))
    return AskResponse(reply=reply, emotion=emotion, score=score, labels=_full_labels(labels))

@app.post("/analyze", response_model=AnalyzeResponse)
async def analyze(
    payload: dict = Body(...),
    request: Request = None,
    _k=Depends(verify_api_key),
    _r=Depends(rate_limit),
):
    """
    後方互換: JSON の "text" でも "prompt" でもOK。Pydantic モデルは使わず dict 受けで 422 を回避。
    """
    raw = (payload.get("text") or payload.get("prompt") or "")
    text = str(raw).strip()
    if not text:
        raise HTTPException(status_code=400, detail="prompt/text が空です。")
    selected_emotion = payload.get("selected_emotion")
    top_emotion, score, labels = resolve_emotion(text, selected_emotion)
    sig = compute_signals(text)
    # 保存（失敗はログのみ）
    try:
        _ = save_stats(labels=labels, signals=sig, top_emotion=top_emotion, score=score, class_id=None)
        ANALYZE_COUNT.labels(emotion=top_emotion).inc()
    except Exception as e:
        rid = getattr(getattr(request, "state", None), "rid", None) or str(uuid.uuid4())
        logger.exception("rid=%s failed to persist /analyze: %s", rid, e)
    return AnalyzeResponse(labels=_full_labels(labels), signals=sig)

# ============== Summary =======================================================
class DayCounts(BaseModel):
    date: str
    counts: Dict[str, int]
    total: int

class SummaryResponse(BaseModel):
    days: int
    daily: List[DayCounts]
    totals: Dict[str, int]
    top_emotion: str

@app.get("/summary", response_model=SummaryResponse)
async def summary(
    days: int = 7,
    class_id: Optional[str] = None,
    tz: str = "Asia/Tokyo",
    include_empty_days: bool = True,
    _k: bool = Depends(verify_api_key),
    _r: bool = Depends(rate_limit),
):
    tzinfo = zoneinfo.ZoneInfo(tz)
    today_local = datetime.now(tzinfo).date()
    start_date_local = today_local - timedelta(days=days - 1)
    from datetime import timezone as _tz
    start_local_aware = datetime.combine(start_date_local, datetime.min.time(), tzinfo=tzinfo)
    start_dt = start_local_aware.astimezone(_tz.utc).replace(tzinfo=None)
    where_clause = [StatsRecord.created_at >= start_dt]
    if class_id:
        where_clause.append(StatsRecord.class_id == class_id)
    with SessionLocal() as db:
        rows = db.execute(select(StatsRecord.created_at, StatsRecord.emotion).where(and_(*where_clause))).all()
    def to_local_date_str(dtobj: datetime) -> str:
        if dtobj.tzinfo is None:
            return dtobj.replace(tzinfo=_tz.utc).astimezone(tzinfo).date().isoformat()
        return dtobj.astimezone(tzinfo).date().isoformat()
    by_day: Dict[str, Dict[str, int]] = {}
    if include_empty_days:
        for i in range(days):
            d = (start_date_local + timedelta(days=i)).isoformat()
            by_day[d] = {k: 0 for k in EMOTION_KEYS}
    for created_at, emo in rows:
        d = to_local_date_str(created_at)
        if d < start_date_local.isoformat() or d > today_local.isoformat():
            continue
        if d not in by_day:
            by_day[d] = {k: 0 for k in EMOTION_KEYS}
        if emo not in by_day[d]:
            by_day[d][emo] = 0
        by_day[d][emo] += 1
    days_sorted = [
        (start_date_local + timedelta(days=i)).isoformat() for i in range(days)
    ] if include_empty_days else sorted(by_day.keys())
    daily: List[DayCounts] = []
    totals = {k: 0 for k in EMOTION_KEYS}
    for d in days_sorted:
        counts = {k: int(by_day.get(d, {}).get(k, 0)) for k in EMOTION_KEYS}
        total = sum(counts.values())
        daily.append(DayCounts(date=d, counts=counts, total=total))
        for k, v in counts.items():
            totals[k] += v
    top_emotion = "中立" if sum(totals.values()) == 0 else max(
        EMOTION_KEYS, key=lambda k: (totals.get(k, 0), -EMOTION_KEYS.index(k))
    )
    return SummaryResponse(days=len(daily), daily=daily, totals=totals, top_emotion=top_emotion)

# ============== Weekly Report ================================================
class WeeklyReportResponse(BaseModel):
    start_date: str
    end_date: str
    tz: str
    days: int
    daily: List[DayCounts]
    totals: Dict[str, int]
    top_emotion: str
    trend: Dict[str, List[str]]
    summary: str
    suggestions: List[str]

REPORT_TTL_SEC = int(os.getenv("NOLOOK_REPORT_TTL_SEC", "10"))

class _TTLCache:
    def __init__(self, ttl: int):
        self.ttl = ttl
        self.store: Dict[tuple, tuple] = {}
    def get(self, key):
        item = self.store.get(key)
        if not item:
            return None
        ts, payload = item
        if (time.time() - ts) < self.ttl:
            return payload
        self.store.pop(key, None)
        return None
    def set(self, key, payload):
        self.store[key] = (time.time(), payload)

_report_cache = _TTLCache(REPORT_TTL_SEC)

@app.get("/weekly_report", response_model=WeeklyReportResponse)
async def weekly_report(
    days: int = 7,
    class_id: Optional[str] = None,
    tz: str = "Asia/Tokyo",
    include_empty_days: bool = True,
    _k: bool = Depends(verify_api_key),
    _r: bool = Depends(rate_limit),
):
    cache_key = (days, class_id or "", tz, include_empty_days)
    cached = _report_cache.get(cache_key)
    if cached:
        try:
            WEEKLY_CACHE_HITS.inc()
        except Exception:
            pass
        return WeeklyReportResponse(**cached)
    try:
        WEEKLY_CACHE_MISSES.inc()
    except Exception:
        pass

    tzinfo = zoneinfo.ZoneInfo(tz)
    today_local = datetime.now(tzinfo).date()
    start_date_local = today_local - timedelta(days=days - 1)
    from datetime import timezone as _tz
    start_local_aware = datetime.combine(start_date_local, datetime.min.time(), tzinfo=tzinfo)
    start_dt = start_local_aware.astimezone(_tz.utc).replace(tzinfo=None)
    where_clause = [StatsRecord.created_at >= start_dt]
    if class_id:
        where_clause.append(StatsRecord.class_id == class_id)
    with SessionLocal() as db:
        rows = db.execute(select(StatsRecord.created_at, StatsRecord.emotion).where(and_(*where_clause))).all()

    def to_local_date_str(dtobj: datetime) -> str:
        if dtobj.tzinfo is None:
            return dtobj.replace(tzinfo=_tz.utc).astimezone(tzinfo).date().isoformat()
        return dtobj.astimezone(tzinfo).date().isoformat()

    by_day: Dict[str, Dict[str, int]] = {}
    if include_empty_days:
        for i in range(days):
            d = (start_date_local + timedelta(days=i)).isoformat()
            by_day[d] = {k: 0 for k in EMOTION_KEYS}

    for created_at, emo in rows:
        d = to_local_date_str(created_at)
        if d < start_date_local.isoformat() or d > today_local.isoformat():
            continue
        if d not in by_day:
            by_day[d] = {k: 0 for k in EMOTION_KEYS}
        if emo not in by_day[d]:
            by_day[d][emo] = 0
        by_day[d][emo] += 1

    days_sorted = [
        (start_date_local + timedelta(days=i)).isoformat() for i in range(days)
    ] if include_empty_days else sorted(by_day.keys())

    daily: List[DayCounts] = []
    totals = {k: 0 for k in EMOTION_KEYS}
    for d in days_sorted:
        counts = {k: int(by_day.get(d, {}).get(k, 0)) for k in EMOTION_KEYS}
        total = sum(counts.values())
        daily.append(DayCounts(date=d, counts=counts, total=total))
        for k, v in counts.items():
            totals[k] += v

    grand_total = sum(totals.values())
    top_emotion = "中立" if grand_total == 0 else max(
        EMOTION_KEYS, key=lambda k: (totals.get(k, 0), -EMOTION_KEYS.index(k))
    )

    split = max(1, days // 2)
    older_days = days_sorted[:split]
    recent_days = days_sorted[split:]

    def sum_range(day_keys: List[str]) -> Dict[str, int]:
        x = {k: 0 for k in EMOTION_KEYS}
        for dd in day_keys:
            for k in EMOTION_KEYS:
                x[k] += int(by_day.get(dd, {}).get(k, 0))
        return x

    older = sum_range(older_days)
    recent = sum_range(recent_days)
    rising = [k for k in EMOTION_KEYS if recent.get(k, 0) > older.get(k, 0)]
    falling = [k for k in EMOTION_KEYS if recent.get(k, 0) < older.get(k, 0)]

    def pct(n: int) -> float:
        return (n / grand_total) if grand_total else 0.0

    p_sad = pct(totals.get("悲しい", 0))
    p_fear = pct(totals.get("不安", 0))
    p_ang = pct(totals.get("怒り", 0))
    p_tired = pct(totals.get("しんどい", 0))

    notes: List[str] = []
    if grand_total > 0:
        peak_day = max(daily, key=lambda r: r.total).date if daily else start_date_local.isoformat()
        notes.append(f"ピーク日: {peak_day}")
        notes.append(f"最多感情: {top_emotion} ({totals.get(top_emotion, 0)}件)")
        if rising:
            notes.append("増加: " + ", ".join(rising))
        if falling:
            notes.append("減少: " + ", ".join(falling))

    # 先生向けのやわらかい提案
    suggestions: List[str] = []
    if p_sad + p_fear >= 0.6 and grand_total >= 5:
        suggestions.append("今週は不安・悲しみがやや多めでした。無理のない声かけで気持ちを受け止め、必要に応じて相談先の案内をそっと添えてください。")
    if p_ang >= 0.3:
        suggestions.append("怒りの表現が目立ちました。安全を最優先に、落ち着ける場をつくってから事実確認とルールの共有を丁寧に。")
    if p_tired >= 0.3:
        suggestions.append("しんどさがやや多めです。休息の確保や課題量の見直し、睡眠リズムの整えなど、できる一歩を一緒に検討できると良さそうです。")
    if not suggestions and grand_total > 0:
        suggestions.append("小さな良い変化を見つけて共有しながら、来週も安心して話せる雰囲気づくりを続けましょう。")

    summary_text = (
        "今週の投稿は少なめでした。来週も負担にならない範囲で、あたたかく見守っていきましょう。" if grand_total == 0 else
        f"合計{grand_total}件でした。いちばん多かったのは『{top_emotion}』。直近は{', '.join(rising) if rising else '大きな増減は'}見られませんでした。"
    )

    result = WeeklyReportResponse(
        start_date=start_date_local.isoformat(),
        end_date=today_local.isoformat(),
        tz=tz,
        days=len(days_sorted),
        daily=daily,
        totals=totals,
        top_emotion=top_emotion,
        trend={"rising": rising, "falling": falling},
        summary=summary_text,
        suggestions=notes + suggestions,
    )
    try:
        payload = result.model_dump() if hasattr(result, "model_dump") else result.dict()
        _report_cache.set(cache_key, payload)
    except Exception:
        pass
    return result

# ============== /metrics ======================================================
@app.get("/metrics")
def metrics(_k: bool = Depends(verify_api_key)):
    data = generate_latest(PROM_REG)
    return StreamingResponse(io.BytesIO(data), media_type=CONTENT_TYPE_LATEST)

# ============== Export (JSON / XLSX) =========================================
def _labels_onehot_json(labels: dict) -> str:
    # 日本語キーで one-hot / 0.0/1.0（小数1桁）を安定出力
    full = _full_labels(labels or {})
    ordered = {k: float(full.get(k, 0.0)) for k in EMOTION_KEYS}
    return "{" + ", ".join([f"\"{k}\": {v:.1f}" for k, v in ordered.items()]) + "}"

def _parse_iso_date(value: str, end_of_day: bool = False) -> dt.datetime:
    try:
        if "T" in value:
            dtobj = dt.datetime.fromisoformat(value.replace("Z", "+00:00"))
            if dtobj.tzinfo is not None:
                dtobj = dtobj.astimezone(dt.timezone.utc).replace(tzinfo=None)
            return dtobj
        d = dt.date.fromisoformat(value)
        if end_of_day:
            return dt.datetime(d.year, d.month, d.day, 23, 59, 59)
        return dt.datetime(d.year, d.month, d.day, 0, 0, 0)
    except Exception:
        raise HTTPException(status_code=422, detail=f"Invalid ISO date/datetime: {value}")

@app.get("/export")
def export_logs(
    format: Literal["xlsx","json"] = Query("xlsx", description="出力形式 xlsx|json"),
    since: Optional[str] = Query(None, description="開始 (YYYY-MM-DD or ISO datetime)"),
    until: Optional[str] = Query(None, description="終了 (YYYY-MM-DD or ISO datetime)"),
    class_id: Optional[str] = Query(default=None, description="クラスIDで絞り込み"),
    tz: str = Query("Asia/Tokyo", description="日付整形に使うタイムゾーン（例: Asia/Tokyo）"),
):
    # デフォルト範囲 = 直近7日間
    today = dt.date.today()
    if not since:
        since = (today - dt.timedelta(days=7)).isoformat()
    if not until:
        until = today.isoformat()

    # 期間
    start = _parse_iso_date(since, end_of_day=False)
    end   = _parse_iso_date(until, end_of_day=True)
    if end < start:
        raise HTTPException(status_code=422, detail="until は since 以上にしてください")

    # データ取得
    with SessionLocal() as db:
        stmt = select(StatsRecord).where(and_(StatsRecord.created_at >= start, StatsRecord.created_at <= end))
        if class_id:
            stmt = stmt.where(StatsRecord.class_id == class_id)
        recs: List[StatsRecord] = db.execute(stmt).scalars().all()

    # 並び順: created_at(昇順) -> id(昇順)
    recs_sorted = sorted(recs, key=lambda r: ((r.created_at or dt.datetime.min), r.id or 0))

    # DEMO（先生向け）
    demo_rows = [{
        "id": r.id,
        "記録日時": _fmt_created_local(r.created_at, tz),
        "クラス": _class_display(r.class_id),
        "感情": r.emotion,
        "トピック": _topics_join(r.topic_tags),
        "信頼度": round(float(r.score), 3),
    } for r in recs_sorted]

    # RAW（解析向け）
    raw_rows = []
    for r in recs_sorted:
        labels_json = _labels_onehot_json(r.labels or {})
        raw_rows.append({
            "id": r.id,
            "created_at": _fmt_created_local(r.created_at, tz),
            "class_id": _class_display(r.class_id),
            "emotion_ja": r.emotion,
            "emotion": (_width_fold(r.emotion) if r.emotion else ""),
            "score": round(float(r.score), 3),
            "relationship_mention": bool(r.relationship_mention),
            "negation_index": float(r.negation_index),
            "avoidance": float(r.avoidance),
            "labels_json": labels_json,
            "topic_tags": _topics_join(r.topic_tags),
        })

    # JSON出力
    if format == "json":
        def iso_z(d: Optional[dt.datetime]) -> Optional[str]:
            if d is None:
                return None
            if d.tzinfo is None:
                d = d.replace(tzinfo=dt.timezone.utc)
            return d.astimezone(dt.timezone.utc).isoformat().replace("+00:00", "Z")

        rows_json = []
        for r in recs_sorted:
            rows_json.append({
                "id": r.id,
                "created_at": iso_z(r.created_at),
                "class_id": r.class_id,
                "emotion": r.emotion,
                "score": float(r.score),
                "relationship_mention": bool(r.relationship_mention),
                "negation_index": float(r.negation_index),
                "avoidance": float(r.avoidance),
                "labels": _full_labels(r.labels or {}),
                "topic_tags": list(r.topic_tags or []),
            })
        try:
            body_preview = json.dumps({"count": len(rows_json)}, ensure_ascii=False).encode("utf-8")
            EXPORT_BYTES.labels(format="json").inc(len(body_preview))
        except Exception:
            pass

        return JSONResponse({
            "rows": rows_json,
            "count": len(rows_json),
            "class_id": class_id,
            "since": start.isoformat(),
            "until": end.isoformat(),
            "tz": tz,
        })

    # XLSX出力
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    def autosize(ws):
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = 0
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(40, max(10, max_len + 2))

    wb = Workbook()

    # DEMO
    ws_demo = wb.active
    ws_demo.title = "DEMO"
    demo_headers = ["id", "記録日時", "クラス", "感情", "トピック", "信頼度"]
    ws_demo.append(demo_headers)
    for row in demo_rows:
        ws_demo.append([row.get(h, "") for h in demo_headers])
    for c in ws_demo[1]:
        c.font = Font(bold=True)
    ws_demo.freeze_panes = "A2"
    ws_demo.auto_filter.ref = ws_demo.dimensions
    autosize(ws_demo)

    # RAW
    ws_raw = wb.create_sheet("RAW")
    raw_headers = ["id","created_at","class_id","emotion_ja","emotion","score",
                   "relationship_mention","negation_index","avoidance",
                   "labels_json","topic_tags"]
    ws_raw.append(raw_headers)
    for row in raw_rows:
        ws_raw.append([row.get(h, "") for h in raw_headers])
    for c in ws_raw[1]:
        c.font = Font(bold=True)
    ws_raw.freeze_panes = "A2"
    ws_raw.auto_filter.ref = ws_raw.dimensions
    autosize(ws_raw)

    bio = io.BytesIO()
    wb.save(bio)
    try:
        size = bio.getbuffer().nbytes
        EXPORT_BYTES.labels(format="xlsx").inc(size)
    except Exception:
        pass
    bio.seek(0)
    filename = f'export_demo+raw_{class_id or "all"}_{start.date()}_{end.date()}.xlsx'
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

# ============== __main__ ======================================================
if __name__ == "__main__":
    # 簡易セルフテスト（RUN_TESTS=1 の時のみ）
    if os.getenv("RUN_TESTS") == "1":
        from fastapi.testclient import TestClient
        client = TestClient(app)

        # /ask
        r = client.post("/ask", json={"prompt": "今日は萎えたー"})
        assert r.status_code == 200, r.text
        assert r.json()["emotion"] in EMOTION_KEYS, r.json()

        # /analyze (text 後方互換)
        r2 = client.post("/analyze", json={"text": "今日はテストで不安"})
        assert r2.status_code == 200
        keys = list(r2.json()["labels"].keys())
        assert keys == EMOTION_KEYS, keys

        # /summary
        r3 = client.get("/summary", params={"days": 7})
        assert r3.status_code == 200

        # API key check
        os.environ["API_KEY"] = "test"
        r4 = client.post("/ask", json={"prompt": "ok"})
        assert r4.status_code == 401

        print("All tests passed")

    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=int(os.getenv("PORT", "8000")))

